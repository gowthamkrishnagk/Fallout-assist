"""
Ingestion pipeline — two sources:
  1. Jira resolved/closed/cancelled SAC tickets via JQL
  2. Uploaded workaround documents (PDF / Word / txt / xlsx)

Both write to the same Chroma collection via vectordb.py.
"""

import hashlib
import json
import os
import re
import threading
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv
load_dotenv()

import embedder
import vectordb
from textclean import clean_text, extract_fix_block

_STATE_FILE  = Path(__file__).parent / "trackers" / "ingest_state.json"
_ingest_lock = threading.Lock()

# Structured fields to extract from ticket descriptions (ORDER_FALLOUT format)
_STRUCTURED_FIELDS = [
    "Order Reason", "Order Type", "Step", "Failed Step",
    "Error Description", "Error Code", "Sub Type", "Order Sub Type",
    "Service Type", "Action",
]


# ── Jira ingestion ────────────────────────────────────────────────────────────

def _get_jira(cfg: dict):
    from jira import JIRA
    email     = os.getenv("JIRA_EMAIL") or cfg["jira"]["email"]
    api_token = os.getenv("JIRA_API_TOKEN") or cfg["jira"]["api_token"]
    url       = cfg["jira"]["url"]
    return JIRA(server=url, basic_auth=(email, api_token))


def _search_jql(jira, jql: str, fields: str, max_issues=None, expand=None) -> list:
    """Atlassian removed the old GET /rest/api/{2,3}/search (HTTP 410, CHANGE-2046).
    This uses the replacement cursor-paginated /rest/api/2/search/jql endpoint and
    wraps the raw results as jira Issue objects, so all downstream code is unchanged.
    The /2/ base is deliberate: it keeps description/comment as plain text (the /3/
    endpoint returns Atlassian Document Format JSON, which would break text parsing).

    Pass expand="changelog" to include each issue's change history (needed to find
    when/who resolved the ticket — see _resolution_event).

    Pagination is by opaque nextPageToken (there is no startAt/total anymore)."""
    from jira.resources import Issue
    server = jira._options["server"].rstrip("/")
    url    = f"{server}/rest/api/2/search/jql"
    issues = []
    token  = None
    while True:
        params = {"jql": jql, "maxResults": 100, "fields": fields}
        if expand:
            params["expand"] = expand
        if token:
            params["nextPageToken"] = token
        resp = jira._session.get(url, params=params)
        resp.raise_for_status()
        data  = resp.json()
        batch = data.get("issues", [])
        for raw in batch:
            issues.append(Issue(jira._options, jira._session, raw=raw))
            if max_issues and len(issues) >= max_issues:
                return issues
        token = data.get("nextPageToken")
        if not token or not batch:
            break
    return issues


def _approx_count(jira, jql: str):
    """Fast count for previews — the new search API no longer returns a total.
    Returns an int, or None if the endpoint is unavailable."""
    try:
        server = jira._options["server"].rstrip("/")
        resp   = jira._session.post(f"{server}/rest/api/2/search/approximate-count",
                                    json={"jql": jql})
        resp.raise_for_status()
        return resp.json().get("count")
    except Exception:
        return None


def _clean(text: str) -> str:
    if not text:
        return ""
    text = re.sub(r'\r\n|\r', '\n', text)
    text = re.sub(r'\n{3,}', '\n\n', text)
    return text.strip()


def _to_epoch(s: str) -> float:
    """Jira datetime ('2026-06-04T19:40:12.345+0530') -> epoch seconds, for
    recency sorting. Returns 0.0 if unparseable."""
    if not s:
        return 0.0
    try:
        # Normalize the '+0530' offset to '+05:30' so fromisoformat accepts it.
        s2 = re.sub(r'([+-]\d{2})(\d{2})$', r'\1:\2', s.strip())
        return datetime.fromisoformat(s2).timestamp()
    except Exception:
        return 0.0


def _clean_for_embed(text: str) -> str:
    """Strip MSISDNs, order numbers, Salesforce IDs, emoji/mojibake and markup —
    keep step/error context. Delegates to the shared canonical cleaner so ingest
    and search normalize identically."""
    return clean_text(text)


def _parse_structured_fields(desc: str) -> dict:
    """{field: value} for the structured labels found in a ticket description.
    Handles Jira wiki markup bold (*Field:*), plain text, and table cells. Pure
    numbers are skipped (they're order/account IDs, not meaningful text)."""
    out: dict = {}
    for field in _STRUCTURED_FIELDS:
        # Matches: *Order Reason:* value  |  Order Reason: value  |  | Order Reason | value |
        m = re.search(
            rf'(?:^|\n|\|)\s*\*?{re.escape(field)}\*?\s*[:\|]\s*\*?([^*\n\|]{{1,200}}?)\*?\s*(?:\||$|\n)',
            desc, re.IGNORECASE | re.MULTILINE
        )
        if m:
            val = m.group(1).strip().strip('*').strip()
            if val and not re.match(r'^\d+$', val):
                out[field] = val
    return out


def _extract_description_fields(issue) -> str:
    """Extract structured fields (Order Reason, Order Type, Step, Error Description…)
    from the ticket description as a joined "Field: value" string for embedding/display."""
    desc   = getattr(issue.fields, 'description', '') or ''
    fields = _parse_structured_fields(desc)
    return '\n'.join(f"{k}: {v}" for k, v in fields.items())


# Authors whose comments are automated noise — never real workarounds
_BOT_AUTHORS = {"sac bot", "automation for jira"}


def _is_bot(display_name: str) -> bool:
    return display_name.lower().strip() in _BOT_AUTHORS


def _extract_error_context(issue) -> str:
    """Pull error description from two sources:
    1. PR_Error_Description in SAC BOT comments (e.g. '14081 | Plan instance already cancelled')
    2. Error Description field in the ticket description (for errors not logged by SAC BOT)
    Both are embedded into every chunk so searching by error text finds the right resolution."""
    # Source 1: SAC BOT PR_Error_Description field
    comments = issue.fields.comment.comments if issue.fields.comment else []
    for c in reversed(comments):
        if c.author.displayName.lower().strip() != "sac bot":
            continue
        if not c.body:
            continue
        for line in c.body.splitlines():
            if "PR_Error_Description" in line:
                parts = line.split(":", 1)
                if len(parts) == 2:
                    # Same canonical cleaner as Source 2 and the query side, so
                    # the identical error never embeds two different ways.
                    err = clean_text(parts[1])
                    if err:
                        return err

    # Source 2: Description field — handles errors SAC BOT doesn't log (e.g. Asurion errors)
    desc = getattr(issue.fields, 'description', '') or ''
    if isinstance(desc, str) and desc.strip() and not desc.strip().startswith('{'):
        for line in desc.splitlines():
            line = line.strip().strip('*').strip()
            # Look for lines that contain "Error" label or look like an error message
            m = re.search(r'Error(?:\s+Description)?\s*[:\|]\s*(.{20,300})', line, re.IGNORECASE)
            if m:
                return clean_text(m.group(1))

    return ""


# Statuses that mean the ticket reached a resolved/terminal state.
_RESOLVED_STATUSES = {"resolved", "closed", "done", "cancelled", "canceled",
                      "complete", "completed"}

# How close (seconds) a comment must be to the resolution event to count as the
# resolution comment. The user's observation: the real fix is commented within a
# couple of minutes of the ticket being resolved.
_RESOLUTION_WINDOW_SECONDS = 120  # ±2 min

# A bare acknowledgement with no resolution content — never the real fix, even from
# the assignee.
_BARE_CLOSER = re.compile(
    r'(?i)^\s*(done|fixed|ok|okay|n/?a|closing|closed|resolved|thanks?|ty|noted)[.! ]*$')

# A routing/meta note (a lead tagging who actually worked the ticket, e.g. "This
# ticket is worked by <name>, the actual shift assignee") — carries no fix, so it
# must never be stored as the resolution. Kept specific so a genuine fix that merely
# mentions reassignment isn't dropped.
_META_NOTE = re.compile(
    r'(?i)(this (ticket|issue) is (being )?(worked|handled) by|shift assignee|actual (shift )?assignee)')


def _resolution_event(issue):
    """From the issue changelog, find WHEN the ticket was resolved and WHO did it.

    Looks for the most recent history entry that either sets a non-empty resolution
    or moves status into a terminal state (_RESOLVED_STATUSES). Taking the *most
    recent* such transition handles reopen→reclose correctly.

    Returns (epoch_seconds, resolver_display_name), or (None, None) when there's no
    changelog (e.g. it wasn't expanded) or no resolving transition."""
    changelog = getattr(issue, "changelog", None)
    histories = getattr(changelog, "histories", None) if changelog else None
    if not histories:
        return None, None

    best_epoch = None
    best_author = None
    for h in histories:
        items = getattr(h, "items", []) or []
        resolved_here = False
        for it in items:
            field  = (getattr(it, "field", "") or "").lower()
            to_str = (getattr(it, "toString", "") or "").strip()
            if field == "resolution" and to_str:
                resolved_here = True
                break
            if field == "status" and to_str.lower() in _RESOLVED_STATUSES:
                resolved_here = True
                break
        if not resolved_here:
            continue
        ep = _to_epoch(getattr(h, "created", "") or "")
        if best_epoch is None or ep >= best_epoch:
            best_epoch  = ep
            author      = getattr(h, "author", None)
            best_author = getattr(author, "displayName", None)
    return best_epoch, best_author


def _get_resolution_comments(issue) -> list[dict]:
    """Return the ticket's actual resolution comment(s) — the work done by the
    ticket's ASSIGNEE, anchored to the moment the ticket was resolved.

    Why assignee-first: a lead often flips the status to Resolved and leaves only a
    meta-note ("This ticket is worked by <name>, the actual shift assignee"), so the
    changelog *resolver* is NOT a reliable source of the real fix. The engineer the
    ticket is assigned to is the one who worked it, so their own comment is the fix.

    Strategy:
      1. _resolution_event → when the ticket was resolved (and who flipped it).
      2. Prefer the ASSIGNEE's substantive comments, anchored to the resolution:
         those within ±2 min of the event, else the single one closest to it.
      3. If the assignee left no comment, fall back to the resolver's comment(s),
         then to any human — same anchoring. No changelog → last human comment.

    Pointer comments ('duplicate, refer to SAC-x') are KEPT — at search time the
    follow-reference logic resolves them to the referenced ticket's real fix.
    When a comment carries a === FIX === block, only that block is stored."""
    comments      = issue.fields.comment.comments if issue.fields.comment else []
    assignee      = getattr(issue.fields, "assignee", None)
    assignee_id   = getattr(assignee, "accountId", None)
    assignee_name = getattr(assignee, "displayName", None)

    def _is_assignee(c) -> bool:
        if assignee_id and getattr(c.author, "accountId", None) == assignee_id:
            return True
        if assignee_name and getattr(c.author, "displayName", None) == assignee_name:
            return True
        return False

    # Build the human-comment pool. A short comment is kept ONLY when it's the
    # assignee's and isn't a bare closer — the assignee's terse note during
    # resolution ("retried", "reprocessed", "re-triggered the step") IS the fix, and
    # the >40-char rule would otherwise discard it. Everyone else needs a substantive
    # body. "worked by <name> / shift assignee" routing notes are dropped outright.
    human = []
    for c in comments:
        if _is_bot(c.author.displayName):
            continue
        raw = (c.body or "").strip()
        if not raw or _META_NOTE.search(raw):
            continue
        is_asg      = _is_assignee(c)
        substantive = len(raw) > 40
        short_asg   = is_asg and len(raw) >= 3 and not _BARE_CLOSER.match(raw)
        if not (substantive or short_asg):
            continue
        human.append({
            "author":      c.author.displayName,
            "body":        extract_fix_block(_clean(c.body)),
            "epoch":       _to_epoch(getattr(c, "created", "") or ""),
            "is_assignee": is_asg,
        })
    if not human:
        return []

    event_epoch, event_author = _resolution_event(issue)

    def _anchor(cands: list) -> list:
        """Comment(s) from `cands` tied to the resolution moment: those within the
        ±window, else the single one closest to it; with no event, the last one."""
        if not cands:
            return []
        if event_epoch:
            within = [c for c in cands
                      if abs(c["epoch"] - event_epoch) <= _RESOLUTION_WINDOW_SECONDS]
            if within:
                within.sort(key=lambda c: c["epoch"])
                return within
            return [min(cands, key=lambda c: abs(c["epoch"] - event_epoch))]
        return cands[-1:]

    # Prefer the ASSIGNEE — the engineer who actually worked the ticket. A lead may
    # have flipped the status and left only a meta-note ("worked by <assignee>"), so
    # the changelog resolver is not a reliable source; the assignee's own comment is.
    assignee_cmts = [c for c in human if c["is_assignee"]]
    if assignee_cmts:
        return _anchor(assignee_cmts)

    # Assignee left no substantive comment → fall back to whoever resolved it, then
    # to any human, anchored to the resolution moment.
    resolver = [c for c in human if event_author and c["author"] == event_author]
    return _anchor(resolver if resolver else human)


def ingest_jira(cfg: dict, progress_cb=None, full: bool = False,
                since_minutes: int | None = None) -> dict:
    """Pull tickets and index them INCREMENTALLY.

    Only NEW or CHANGED tickets are re-embedded: a ticket whose Jira `updated`
    timestamp matches what we stored on the last run is skipped entirely — no
    re-embedding, no DB writes. Tickets that have fallen out of the JQL since the
    last run (e.g. reopened) are pruned from the index. Pass full=True to force a
    complete rebuild (e.g. after changing the embedding model or the format).

    since_minutes: when set (and not full), only FETCH tickets updated within the
    last N minutes (adds `updated >= -Nm` to the JQL) — used by the background
    scheduler so a frequent run pulls just the latest tickets instead of all of
    them. Pruning is skipped in this windowed mode (we only saw a slice).

    Returns {ok, indexed, skipped, up_to_date, pruned, tickets_indexed, error}.
    """
    with _ingest_lock:
        wf          = cfg["workaround_finder"]
        windowed    = (not full) and bool(since_minutes) and since_minutes > 0
        base_jql    = wf["ingest_jql"]
        if windowed:
            base_jql += f" AND updated >= -{int(since_minutes)}m"
        jql         = base_jql + " ORDER BY updated DESC"
        index_path  = str(Path(__file__).parent / wf["index_path"])
        embed_model = cfg["embed"]["model"]

        try:
            jira   = _get_jira(cfg)
            issues = _search_jql(jira, jql,
                                 "summary,description,status,comment,resolution,assignee,updated",
                                 expand="changelog")
            total  = len(issues)
            if progress_cb:
                progress_cb(f"Fetched {total} tickets from Jira")

            # Per-ticket sync state: {jira_key: last-seen `updated` timestamp}.
            # full=True ignores it so everything is rebuilt from scratch.
            state        = _load_state()
            known        = {} if full else dict(state.get("tickets", {}))
            fetched_keys = {issue.key for issue in issues}

            # Full rebuild → recreate the ticket collections from scratch (docs
            # untouched) so stale chunks can't linger AND the collections get
            # correct HNSW search_ef instead of the small legacy default.
            # NOTE: a full rebuild does NOT wipe here. We recreate the collections
            # only just before writing the freshly-embedded chunks (below), so a
            # failure during fetch/embed can never leave the index empty.

            # Prune tickets we indexed before but that no longer match the JQL.
            # Only in full-scope mode — a windowed fetch sees just a recent slice,
            # so absence there does NOT mean the ticket dropped out of the JQL.
            pruned = 0
            if not windowed:
                for key in [k for k in known if k not in fetched_keys]:
                    vectordb.delete_ticket_by_source(key, index_path)
                    known.pop(key, None)
                    pruned += 1

            # Keep only new / changed tickets — skip ones unchanged since last run.
            changed    = []
            up_to_date = 0
            for issue in issues:
                updated = str(getattr(issue.fields, "updated", "") or "")
                if not full and known.get(issue.key) == updated:
                    up_to_date += 1
                    continue
                changed.append((issue, updated))

            if progress_cb:
                progress_cb(f"{len(changed)} new/changed, {up_to_date} unchanged, "
                            f"{pruned} pruned")

            # Upsert (incremental only): clear changed tickets' chunks before
            # re-adding so we never duplicate. A full rebuild recreates the whole
            # collection set below instead, so it skips this.
            if not full:
                for issue, _ in changed:
                    vectordb.delete_ticket_by_source(issue.key, index_path)

            ids = []; step_texts = []; error_texts = []; display_texts = []; metas = []
            skipped = 0
            reindexed = 0

            for i, (issue, updated) in enumerate(changed):
                assignee      = getattr(issue.fields, "assignee", None)
                assignee_name = getattr(assignee, "displayName", "") or ""
                summary       = issue.fields.summary
                url           = f"{jira.server_url}/browse/{issue.key}"
                comments      = _get_resolution_comments(issue)

                # Record the timestamp regardless, so a ticket without a usable
                # resolution comment isn't re-fetched and re-skipped every run.
                known[issue.key] = updated

                if not comments:
                    skipped += 1
                    continue

                reindexed += 1
                error_ctx    = _extract_error_context(issue)
                clean_sum    = _clean_for_embed(summary)
                struct       = _parse_structured_fields(getattr(issue.fields, 'description', '') or '')
                desc_fields  = '\n'.join(f"{k}: {v}" for k, v in struct.items())
                order_type   = struct.get("Order Type", "")
                order_reason = struct.get("Order Reason", "")

                # Extract step name from cleaned summary for focused embedding
                step_match = re.search(r'Failed\s+Step:\s*([^\n]{5,80}?)(?:\s*$)', clean_sum, re.IGNORECASE)
                step_name  = clean_text(step_match.group(1)) if step_match else ''

                for j, c in enumerate(comments):
                    chunk_id = f"{issue.key}_c{j}"

                    # ── Step embed: step name + desc fields (50% of score) ──
                    step_parts = [f"Failed Step: {step_name}"] if step_name else [clean_sum]
                    if desc_fields:
                        step_parts.append(desc_fields)
                    step_text = '\n'.join(step_parts)[:500]

                    # ── Error embed: error description only (50% of score) ──
                    # None if no error context — chunk won't be in error collection
                    error_text = f"Error: {error_ctx}"[:500] if error_ctx else None

                    # ── Display text: full context stored for showing to user ──
                    error_line = f"Error: {error_ctx}\n" if error_ctx else ""
                    desc_line  = f"{desc_fields}\n" if desc_fields else ""
                    display_text = (
                        f"Ticket: {clean_sum}\n"
                        f"{desc_line}"
                        f"{error_line}"
                        f"Assignee: {assignee_name}\n"
                        f"Resolution by {c['author']}:\n{c['body']}"
                    )[:2000]

                    ids.append(chunk_id)
                    step_texts.append(step_text)
                    error_texts.append(error_text)
                    display_texts.append(display_text)
                    metas.append({
                        "source":         "ticket",
                        "source_id":      issue.key,
                        "key":            issue.key,
                        "summary":        summary[:200],
                        "step":           step_name,
                        "error":          error_ctx[:200] if error_ctx else "",
                        "description":    desc_fields[:500] if desc_fields else "",
                        "order_type":     order_type,
                        "order_reason":   order_reason,
                        "status":         issue.fields.status.name,
                        "url":            url,
                        "assignee":       assignee_name,
                        "comment_author": c["author"],
                        "comment_body":   c["body"][:1000],
                        "is_assignee":    str(c["is_assignee"]),
                        "comment_index":  j,
                        "updated_ts":     _to_epoch(updated),   # recency tie-break
                    })

                if progress_cb and i % 20 == 0:
                    progress_cb(f"Processing ticket {i+1}/{len(changed)}")

            # Embed and store only the new/changed chunks (the unchanged tickets'
            # chunks are already in the index, untouched). No new chunks → the run
            # was a no-op refresh, which is a normal success, not an error.
            if ids:
                if progress_cb:
                    progress_cb(f"Embedding {len(ids)} chunks — step + error separately...")
                step_embs  = embedder.embed(step_texts, embed_model)
                # Embed only non-None error texts; keep None positions as None
                error_texts_clean = [t if t else "" for t in error_texts]
                error_embs_all    = embedder.embed(error_texts_clean, embed_model)
                error_embs = [emb if error_texts[k] else None
                              for k, emb in enumerate(error_embs_all)]

                if progress_cb:
                    progress_cb("Storing fresh ticket chunks...")
                # Full rebuild: wipe + recreate the collections HERE — only now that
                # the new chunks are embedded and ready — so the empty window is a
                # split second and a mid-run failure never leaves the KB empty.
                if full:
                    vectordb.reset_ticket_collections(index_path)
                vectordb.add_tickets(ids, step_embs, error_embs, display_texts, metas, index_path)

            # Refresh the BM25 keyword index + ticket graph so hybrid and graph
            # expansion reflect this pass.
            try:
                import retrieval
                retrieval.build_keyword_index(index_path)
            except Exception as _e:
                print(f"[BM25] keyword index refresh skipped: {_e}")
            try:
                import graph
                graph.build_graph(index_path)
            except Exception as _e:
                print(f"[GRAPH] graph refresh skipped: {_e}")

            _save_state({"last_jira_sync": datetime.utcnow().isoformat(),
                         "tickets":      known,
                         "jira_count":   len(known),
                         "ticket_count": total})
            return {"ok": True, "indexed": len(ids), "skipped": skipped,
                    "up_to_date": up_to_date, "pruned": pruned,
                    "tickets_indexed": reindexed}
        except Exception as e:
            return {"ok": False, "error": str(e)}


# ── Document ingestion ────────────────────────────────────────────────────────

def _parse_doc(path: Path) -> str:
    ext = path.suffix.lower()
    if ext == ".txt" or ext == ".md":
        return path.read_text(encoding="utf-8", errors="ignore")
    if ext == ".pdf":
        from pypdf import PdfReader
        reader = PdfReader(str(path))
        return "\n".join(p.extract_text() or "" for p in reader.pages)
    if ext in (".docx", ".doc"):
        from docx import Document
        doc   = Document(str(path))
        lines = []
        for p in doc.paragraphs:
            t = p.text.strip()
            if not t:
                continue
            # Word renders list numbers/bullets outside p.text, so they're lost on
            # extract. Re-add a marker for list-style paragraphs (unless the text
            # already carries one) so step-by-step instructions stay readable.
            style = (p.style.name or "").lower() if p.style is not None else ""
            if "list" in style and not re.match(r'^\s*([-*•‣●]|\d+[.)])', t):
                t = f"• {t}"
            lines.append(t)
        return "\n".join(lines)
    if ext in (".xlsx", ".xls"):
        from openpyxl import load_workbook
        wb   = load_workbook(str(path), data_only=True)
        rows = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                line = " | ".join(str(c) for c in row if c is not None)
                if line.strip():
                    rows.append(line)
        return "\n".join(rows)
    raise ValueError(f"Unsupported file type: {ext}")


def _chunk_text(text: str, chunk_size: int = 800, overlap: int = 100) -> list[str]:
    """Chunk into ~chunk_size-word windows while PRESERVING line breaks, so a
    document's step-by-step / list structure survives into the stored text and
    renders properly (the UI uses white-space: pre-wrap)."""
    lines  = text.splitlines()
    chunks = []
    cur, cur_words = [], 0
    for line in lines:
        w = len(line.split())
        if cur_words + w > chunk_size and cur:
            chunks.append("\n".join(cur))
            # Carry the last ~overlap words (whole lines) into the next chunk.
            keep, kept = [], 0
            for l in reversed(cur):
                lw = len(l.split())
                if kept + lw > overlap:
                    break
                keep.insert(0, l)
                kept += lw
            cur, cur_words = keep, kept
        cur.append(line)
        cur_words += w
    if cur:
        chunks.append("\n".join(cur))
    return [c for c in chunks if len(c.strip()) > 50]


def _extract_doc_step_error(text: str) -> tuple[str, str]:
    """Pull the failed-step and error lines out of a workaround document so it is
    matched on the same (step, error) basis as Jira tickets. Tolerates both
    'failed step- ...' / 'Step: ...' and 'error response- ...' / 'N | ...' forms."""
    step = error = ""
    m = re.search(r'(?:failed\s+)?step\s*[-:]\s*([^\n]{3,120})', text, re.IGNORECASE)
    if m:
        step = m.group(1).strip()
    m = re.search(r'error(?:\s+response|\s+description)?\s*[-:]\s*([^\n]{3,200})',
                  text, re.IGNORECASE)
    if m:
        error = m.group(1).strip()
    if not error:
        m = re.search(r'(\d{1,6}\s*\|[^\n]{3,150})', text)
        if m:
            error = m.group(1).strip()
    # Drop volatile identifiers so they don't dilute the semantic match.
    error = re.sub(r'\(?\bExternalId\s*=\s*[^)\n]*\)?', '', error, flags=re.IGNORECASE)
    error = re.sub(r'\b\d{6,}\b', '', error).strip(' -|')
    step  = re.sub(r'\b\d{6,}\b', '', step).strip(' -|')
    return step, error


def ingest_document(file_bytes: bytes, filename: str, cfg: dict) -> dict:
    """Parse a document, extract its failed step + error, and index it on that
    (step, error) basis — the same as tickets — so it only matches a query whose
    step/error are semantically close, not one that merely shares boilerplate.
    Re-uploading the same filename replaces the previous version."""
    wf          = cfg["workaround_finder"]
    index_path  = str(Path(__file__).parent / wf["index_path"])
    docs_path   = Path(__file__).parent / wf["docs_path"]
    embed_model = cfg["embed"]["model"]

    docs_path.mkdir(parents=True, exist_ok=True)
    dest   = docs_path / filename
    doc_id = hashlib.md5(filename.encode()).hexdigest()[:12]

    # Remove old version's vectors before indexing new ones
    existing = _load_doc_meta()
    if doc_id in existing:
        vectordb.delete_docs_by_source(doc_id, index_path)

    dest.write_bytes(file_bytes)

    try:
        text = _parse_doc(dest)
        if len(text.strip()) < 50:
            return {"ok": False, "error": "No text extracted from document"}

        step, error = _extract_doc_step_error(text)
        # A workaround doc MUST declare its failed step and error — that's what it
        # gets matched on. Reject anything missing either so it can never surface
        # as an unrelated match later.
        if not step or not error:
            dest.unlink(missing_ok=True)
            missing = " and ".join(
                lbl for lbl, val in (("a 'failed step' line", step),
                                     ("an 'error' line", error)) if not val)
            return {"ok": False,
                    "error": f"Document must contain {missing}. "
                             "Add lines like 'failed step- <name>' and "
                             "'error response- <code | message>', then re-upload."}

        # Frame step/error exactly like tickets/queries so vectors share one space.
        step_text  = f"Failed Step: {step}"
        error_text = f"Error: {error}"

        # Full text (newlines preserved) is the display body shown to the user.
        display    = text[:4000]
        meta       = {"source": "doc", "source_id": doc_id, "filename": filename,
                      "chunk": 0, "step": step, "error": error}
        step_embs  = [embedder.embed_one(step_text,  embed_model)]
        error_embs = [embedder.embed_one(error_text, embed_model)]
        vectordb.add_docs_dual([doc_id], step_embs, error_embs, [display], [meta], index_path)
        try:
            import retrieval
            retrieval.build_keyword_index(index_path)
        except Exception as _e:
            print(f"[BM25] keyword index refresh skipped: {_e}")

        _save_doc_meta(doc_id, filename, 1)
        return {"ok": True, "doc_id": doc_id, "chunks": 1, "filename": filename,
                "step": step, "error": error}
    except Exception as e:
        dest.unlink(missing_ok=True)
        return {"ok": False, "error": str(e)}


def delete_document(doc_id: str, cfg: dict) -> dict:
    index_path = str(Path(__file__).parent / cfg["workaround_finder"]["index_path"])
    deleted    = vectordb.delete_docs_by_source(doc_id, index_path)
    docs_path  = Path(__file__).parent / cfg["workaround_finder"]["docs_path"]
    meta       = _load_doc_meta()
    entry      = meta.pop(doc_id, {})
    filename   = entry.get("filename", "")
    if filename:
        (docs_path / filename).unlink(missing_ok=True)
    _save_doc_meta_raw(meta)
    return {"ok": True, "deleted_chunks": deleted, "filename": filename}


def list_documents(cfg: dict) -> list[dict]:
    return list(_load_doc_meta().values())


def ingest_suggestion(step: str, error: str, body: str, sid: str, cfg: dict) -> dict:
    """Index an APPROVED user-submitted workaround on its failure's (step, error) —
    the same dual-embedding basis as tickets/docs — so it surfaces for that failure
    and semantically similar ones. Stored in the DOC collections (which survive a
    ticket re-ingest), but NOT tracked in the document meta or written to disk, so it
    never appears in the Documents UI. suggestions.json is its source of truth."""
    wf          = cfg["workaround_finder"]
    index_path  = str(Path(__file__).parent / wf["index_path"])
    embed_model = cfg["embed"]["model"]

    step, error, body = (step or "").strip(), (error or "").strip(), (body or "").strip()
    if not body:
        return {"ok": False, "error": "empty suggestion body"}
    if not step and not error:
        return {"ok": False, "error": "suggestion has no step/error to index against"}

    chunk_id = f"usr_{sid}"
    # Remove any prior version (re-approval / re-index) before adding.
    vectordb.delete_docs_by_source(chunk_id, index_path)

    # Frame step/error exactly like tickets/queries so vectors share one space. Either
    # may be absent → that side simply isn't embedded (search_dual handles one-sided).
    step_embs  = [embedder.embed_one(f"Failed Step: {step}", embed_model)] if step  else [None]
    error_embs = [embedder.embed_one(f"Error: {error}",      embed_model)] if error else [None]
    meta = {"source": "doc", "source_id": chunk_id, "filename": "💡 User-submitted fix",
            "chunk": 0, "step": step, "error": error, "kind": "user_fix"}
    vectordb.add_docs_dual([chunk_id], step_embs, error_embs, [body[:4000]], [meta], index_path)

    try:
        import retrieval
        retrieval.build_keyword_index(index_path)
    except Exception as _e:
        print(f"[BM25] keyword index refresh skipped: {_e}")

    return {"ok": True, "indexed_id": chunk_id}


# ── Jira ticket preview ───────────────────────────────────────────────────────

def _ticket_to_text(issue) -> str:
    """Build a clean search query from a fetched ticket.
    Extracts structured fields (category, step, description fields, error) —
    strips MSISDNs, order numbers, and Salesforce IDs so they don't dilute the match."""
    summary = issue.fields.summary or ''

    # Category prefix: ORDER_FALLOUT, PORT_OUT, SUSPEND, etc.
    cat_match = re.match(r'^([A-Z_]+)\s*:', summary)
    category  = cat_match.group(1).strip() if cat_match else ''

    # Step from summary: "… - Failed Step: Create Commercial Order Aria - …"
    step_match = re.search(r'Failed\s+Step:\s*([^-\n]+)', summary, re.IGNORECASE)
    step       = step_match.group(1).strip() if step_match else ''

    # Structured fields from description body
    desc_fields = _extract_description_fields(issue)

    # Error code from SAC BOT comment
    error_ctx = _extract_error_context(issue)

    parts = []
    if category:
        parts.append(category)
    if step:
        parts.append(f"Failed Step: {step}")
    if desc_fields:
        parts.append(desc_fields)
    if error_ctx:
        parts.append(f"Error: {error_ctx}")

    # Fallback: cleaned summary if we could extract nothing structured
    return '\n'.join(parts) if parts else _clean_for_embed(summary)


def preview_jql(jql: str, cfg: dict) -> dict:
    try:
        jira   = _get_jira(cfg)
        sample = _search_jql(jira, jql + " ORDER BY updated DESC", "summary,status", max_issues=5)
        count  = _approx_count(jira, jql)
        return {
            "ok":    True,
            "count": count if count is not None else len(sample),
            "sample": [{"key": i.key, "summary": i.fields.summary,
                        "status": i.fields.status.name}
                       for i in sample],
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}


# Small in-process cache for live duplicate-follow fetches — resolutions rarely
# change, and the same referenced ticket can be hit across queries.
_resolution_cache: dict = {}


def fetch_ticket_resolution(ticket_id: str, cfg: dict) -> dict:
    """Live-fetch a ticket from Jira and return its best resolution comment, to
    follow a 'duplicate, refer to SAC-x' pointer to the real fix when SAC-x isn't
    indexed. Bots are excluded and the comment is anchored to the resolution event by
    _get_resolution_comments. Returns {key, url, comment, author} or {} if the ticket
    can't be fetched, has no usable resolution, or is itself a pointer."""
    if ticket_id in _resolution_cache:
        return _resolution_cache[ticket_id]
    try:
        jira     = _get_jira(cfg)
        issue    = jira.issue(ticket_id, fields="summary,comment,assignee",
                              expand="changelog")
        comments = _get_resolution_comments(issue)
        if not comments:
            return {}
        c      = comments[-1]   # the resolution comment (latest within the window)
        result = {
            "key":     issue.key,
            "url":     f"{jira.server_url}/browse/{issue.key}",
            "comment": c["body"],
            "author":  c["author"],
        }
        _resolution_cache[ticket_id] = result
        return result
    except Exception as e:
        print(f"[DUP-FOLLOW] could not fetch {ticket_id}: {str(e)[:100]}")
        return {}


def fetch_ticket_text(ticket_id: str, cfg: dict) -> str:
    index_path = str(Path(__file__).parent / cfg["workaround_finder"]["index_path"])
    try:
        jira  = _get_jira(cfg)
        issue = jira.issue(ticket_id, fields="summary,description,comment,assignee")
        return _ticket_to_text(issue)
    except Exception:
        # Live Jira fetch failed (not found / no permission / auth). If the ticket
        # is already in our index, build the query from its stored step + error so
        # the search still works without Jira.
        meta  = vectordb.get_ticket_meta(ticket_id, index_path)
        parts = []
        if meta.get("step"):
            parts.append(f"Failed Step: {meta['step']}")
        if meta.get("description"):     # carries Order Reason / Order Type labels
            parts.append(meta["description"])
        if meta.get("error"):
            parts.append(f"Error: {meta['error']}")
        if parts:
            return "\n".join(parts)
        raise   # not indexed either → surface the original error to the caller


# ── State helpers ─────────────────────────────────────────────────────────────

def get_status(cfg: dict) -> dict:
    state     = _load_state()
    index_path = str(Path(__file__).parent / cfg["workaround_finder"]["index_path"])
    return {
        "total_chunks":    vectordb.count(index_path),
        "last_jira_sync":  state.get("last_jira_sync"),
        "jira_count":      state.get("jira_count", 0),
        "doc_count":       len(_load_doc_meta()),
    }


def _load_state() -> dict:
    try:
        return json.loads(_STATE_FILE.read_text())
    except Exception:
        return {}


def _save_state(update: dict):
    state = _load_state()
    state.update(update)
    _STATE_FILE.parent.mkdir(parents=True, exist_ok=True)
    _STATE_FILE.write_text(json.dumps(state, indent=2))


_DOC_META_FILE = Path(__file__).parent / "trackers" / "doc_meta.json"


def _load_doc_meta() -> dict:
    try:
        return json.loads(_DOC_META_FILE.read_text())
    except Exception:
        return {}


def _save_doc_meta(doc_id: str, filename: str, chunks: int):
    meta = _load_doc_meta()
    meta[doc_id] = {"doc_id": doc_id, "filename": filename, "chunks": chunks,
                    "indexed_at": datetime.utcnow().isoformat()}
    _save_doc_meta_raw(meta)


def _save_doc_meta_raw(meta: dict):
    _DOC_META_FILE.parent.mkdir(parents=True, exist_ok=True)
    _DOC_META_FILE.write_text(json.dumps(meta, indent=2))
